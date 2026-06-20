"""
Backoffice - Ventas Telefónicas Locales
Aplicación de escritorio (tkinter) para administrar sucursales,
clientes, usuarios y sincronizar con GitHub Pages.
"""
import csv
import json
import os
import re
import sys
import threading
import tkinter as tk
from datetime import date
from tkinter import filedialog, messagebox, simpledialog, ttk

# Asegurar que el directorio actual sea el del script
if getattr(sys, 'frozen', False):
    _BASE_DIR = os.path.dirname(sys.executable)
else:
    _BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    sys.path.insert(0, _BASE_DIR)
os.chdir(_BASE_DIR)

from database import get_conn, hash_password, init_db
from excel_import import import_excel
from github_sync import GitHubSync

CONFIG_FILE = os.path.join(_BASE_DIR, 'config.json')


# ─── Configuración local ──────────────────────────────────────────────────────

def load_config() -> dict:
    defaults = {
        'github_token':  '',
        'github_user':   'WALLE802',
        'github_repo':   'VENTAS_CLIENTES',
        'github_branch': 'main',
        'github_path_prefix': '',
        'git_repo_path': '',
    }
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, encoding='utf-8') as f:
            return {**defaults, **json.load(f)}
    return defaults


def save_config(cfg: dict) -> None:
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


# ─── Aplicación principal ─────────────────────────────────────────────────────

class BackofficeApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Backoffice · Ventas Telefónicas")
        self.geometry("960x680")
        self.minsize(800, 550)

        init_db()
        self.cfg = load_config()
        self._branch_combos: list[ttk.Combobox] = []

        self._setup_style()
        self._build_header()
        self._build_notebook()

    # ─── Estilos ──────────────────────────────────────────────────────────────

    def _setup_style(self):
        style = ttk.Style(self)
        style.theme_use('clam')
        style.configure('TNotebook.Tab', padding=[12, 5], font=('Segoe UI', 10))
        style.configure('Treeview',        rowheight=26, font=('Segoe UI', 9))
        style.configure('Treeview.Heading', font=('Segoe UI', 9, 'bold'))
        style.configure('TButton',          padding=[8, 5])

    def _build_header(self):
        bar = tk.Frame(self, bg='#1a73e8', height=48)
        bar.pack(fill='x')
        bar.pack_propagate(False)
        tk.Label(
            bar, text="📞  Backoffice · Ventas Telefónicas",
            bg='#1a73e8', fg='white', font=('Segoe UI', 13, 'bold')
        ).pack(side='left', padx=16, pady=10)
        tk.Button(
            bar, text="🌐  Ver página web",
            bg='#1557b0', fg='white', font=('Segoe UI', 9, 'bold'),
            relief='flat', cursor='hand2', activebackground='#0d47a1',
            activeforeground='white', bd=0, padx=10,
            command=self._open_website
        ).pack(side='right', padx=16, pady=10)

    def _open_website(self):
        import webbrowser
        user = self.cfg.get('github_user', 'WALLE802')
        repo = self.cfg.get('github_repo', 'VENTAS_CLIENTES')
        url = f"https://{user.lower()}.github.io/{repo}/"
        webbrowser.open(url)

    def _build_notebook(self):
        nb = ttk.Notebook(self)
        nb.pack(fill='both', expand=True, padx=6, pady=6)

        self._build_tab_sucursales(nb)
        self._build_tab_clientes(nb)
        self._build_tab_usuarios(nb)
        self._build_tab_sincronizar(nb)
        self._build_tab_mensaje(nb)
        self._build_tab_reportes(nb)

        # Poblar los comboboxes de sucursal ahora que todos los tabs están construidos
        self._sync_branch_combos()

    # ─── TAB: SUCURSALES ──────────────────────────────────────────────────────

    def _build_tab_sucursales(self, nb):
        tab = ttk.Frame(nb)
        nb.add(tab, text='🏢  Sucursales')

        left = ttk.LabelFrame(tab, text='Sucursales activas')
        left.pack(side='left', fill='both', expand=True, padx=10, pady=10)

        self.branch_lb = tk.Listbox(left, font=('Segoe UI', 12),
                                    selectbackground='#1a73e8', activestyle='none')
        sb = ttk.Scrollbar(left, command=self.branch_lb.yview)
        self.branch_lb.configure(yscrollcommand=sb.set)
        self.branch_lb.pack(side='left', fill='both', expand=True, padx=(5,0), pady=5)
        sb.pack(side='right', fill='y', pady=5)

        right = ttk.LabelFrame(tab, text='Administrar')
        right.pack(side='right', fill='y', padx=10, pady=10)

        ttk.Label(right, text='Nombre:').pack(padx=12, pady=(14,2), anchor='w')
        self.branch_entry = ttk.Entry(right, width=22)
        self.branch_entry.pack(padx=12, pady=2)
        self.branch_entry.bind('<Return>', lambda _: self._add_branch())

        ttk.Button(right, text='➕  Agregar',  command=self._add_branch).pack(padx=12, pady=6, fill='x')
        ttk.Button(right, text='🗑  Eliminar', command=self._del_branch).pack(padx=12, pady=6, fill='x')

        self._refresh_branches()

    def _refresh_branches(self):
        self.branch_lb.delete(0, 'end')
        conn = get_conn()
        rows = conn.execute("SELECT name FROM branches ORDER BY name").fetchall()
        conn.close()
        for r in rows:
            self.branch_lb.insert('end', r['name'])
        self._sync_branch_combos()

    def _sync_branch_combos(self):
        conn = get_conn()
        values = [r['name'] for r in conn.execute(
            "SELECT name FROM branches ORDER BY name"
        ).fetchall()]
        conn.close()
        for cb in self._branch_combos:
            current = cb.get()
            cb['values'] = values
            if current in values:
                cb.set(current)

    def _add_branch(self):
        name = self.branch_entry.get().strip()
        if not name:
            messagebox.showwarning("Atención", "Ingresá el nombre de la sucursal.", parent=self)
            return
        conn = get_conn()
        try:
            conn.execute("INSERT INTO branches (name) VALUES (?)", (name,))
            conn.commit()
            self.branch_entry.delete(0, 'end')
            self._refresh_branches()
        except Exception:
            messagebox.showerror("Error", f"Ya existe una sucursal llamada '{name}'.", parent=self)
        finally:
            conn.close()

    def _del_branch(self):
        sel = self.branch_lb.curselection()
        if not sel:
            messagebox.showwarning("Atención", "Seleccioná una sucursal.", parent=self)
            return
        name = self.branch_lb.get(sel[0])
        if not messagebox.askyesno(
            "Confirmar", f"¿Eliminar la sucursal '{name}' y todos sus clientes?\nEsta acción no se puede deshacer.",
            parent=self
        ):
            return
        conn = get_conn()
        conn.execute("DELETE FROM clients  WHERE branch = ?", (name,))
        conn.execute("DELETE FROM branches WHERE name   = ?", (name,))
        conn.commit()
        conn.close()
        self._refresh_branches()

    # ─── TAB: CLIENTES ────────────────────────────────────────────────────────

    def _build_tab_clientes(self, nb):
        tab = ttk.Frame(nb)
        nb.add(tab, text='👥  Clientes')

        # Barra superior
        top = ttk.Frame(tab)
        top.pack(fill='x', padx=10, pady=6)

        ttk.Label(top, text='Sucursal:').pack(side='left')
        self.cl_branch = ttk.Combobox(top, state='readonly', width=16)
        self.cl_branch.pack(side='left', padx=(4, 10))
        self.cl_branch.bind('<<ComboboxSelected>>', lambda _: self._load_clients())
        self._branch_combos.append(self.cl_branch)

        ttk.Button(top, text='📂 Importar Excel',    command=self._import_excel).pack(side='left', padx=3)
        ttk.Button(top, text='🗑 Limpiar sucursal',  command=self._clear_clients).pack(side='left', padx=3)
        ttk.Button(top, text='❌ Eliminar fila',      command=self._del_client).pack(side='left', padx=3)

        # Tabla
        frame = ttk.Frame(tab)
        frame.pack(fill='both', expand=True, padx=10, pady=(0, 5))

        cols = ('DNI', 'Nombre', 'Teléfono', 'Tel.2', 'Tel.3', 'Última Compra')
        self.cl_tree = ttk.Treeview(frame, columns=cols, show='headings')
        widths = (90, 200, 110, 110, 110, 120)
        for col, w in zip(cols, widths):
            self.cl_tree.heading(col, text=col)
            self.cl_tree.column(col, width=w, minwidth=60)

        vsb = ttk.Scrollbar(frame, orient='vertical',   command=self.cl_tree.yview)
        hsb = ttk.Scrollbar(frame, orient='horizontal', command=self.cl_tree.xview)
        self.cl_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.cl_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        self.cl_status = ttk.Label(tab, text='')
        self.cl_status.pack(pady=2)

    def _load_clients(self):
        branch = self.cl_branch.get()
        if not branch:
            return
        self.cl_tree.delete(*self.cl_tree.get_children())
        conn = get_conn()
        rows = conn.execute(
            "SELECT id, dni, nombre, telefono, tel2, tel3, ultima_compra "
            "FROM clients WHERE branch = ? ORDER BY nombre",
            (branch,)
        ).fetchall()
        conn.close()
        for r in rows:
            self.cl_tree.insert('', 'end', iid=str(r['id']),
                                values=(r['dni'], r['nombre'], r['telefono'],
                                        r['tel2'], r['tel3'], r['ultima_compra']))
        self.cl_status.config(text=f"{len(rows)} cliente(s) en '{branch}'")

    def _import_excel(self):
        branch = self.cl_branch.get()
        if not branch:
            messagebox.showwarning("Atención", "Seleccioná una sucursal primero.", parent=self)
            return
        path = filedialog.askopenfilename(
            title='Seleccionar archivo Excel',
            filetypes=[('Excel', '*.xlsx *.xls'), ('Todos', '*.*')],
            parent=self
        )
        if not path:
            return
        try:
            n = import_excel(path, branch)
            messagebox.showinfo("Listo", f"✅ {n} cliente(s) importados a '{branch}'.", parent=self)
            self._load_clients()
        except Exception as e:
            messagebox.showerror("Error al importar", str(e), parent=self)

    def _clear_clients(self):
        branch = self.cl_branch.get()
        if not branch:
            messagebox.showwarning("Atención", "Seleccioná una sucursal.", parent=self)
            return
        if not messagebox.askyesno(
            "Confirmar", f"¿Eliminar TODOS los clientes de '{branch}'?", parent=self
        ):
            return
        conn = get_conn()
        conn.execute("DELETE FROM clients WHERE branch = ?", (branch,))
        conn.commit()
        conn.close()
        self._load_clients()

    def _del_client(self):
        sel = self.cl_tree.selection()
        if not sel:
            messagebox.showwarning("Atención", "Seleccioná un cliente.", parent=self)
            return
        if not messagebox.askyesno("Confirmar", "¿Eliminar el cliente seleccionado?", parent=self):
            return
        conn = get_conn()
        conn.execute("DELETE FROM clients WHERE id = ?", (sel[0],))
        conn.commit()
        conn.close()
        self._load_clients()

    # ─── TAB: USUARIOS ────────────────────────────────────────────────────────

    def _build_tab_usuarios(self, nb):
        tab = ttk.Frame(nb)
        nb.add(tab, text='👤  Usuarios')

        # Tabla izquierda
        left = ttk.Frame(tab)
        left.pack(side='left', fill='both', expand=True, padx=10, pady=10)

        cols = ('Usuario', 'Sucursal')
        self.usr_tree = ttk.Treeview(left, columns=cols, show='headings')
        for col in cols:
            self.usr_tree.heading(col, text=col)
            self.usr_tree.column(col, width=160)
        vsb = ttk.Scrollbar(left, command=self.usr_tree.yview)
        self.usr_tree.configure(yscrollcommand=vsb.set)
        self.usr_tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')
        self.usr_tree.bind('<<TreeviewSelect>>', self._on_user_select)

        # Formulario derecho
        right = ttk.LabelFrame(tab, text='Datos del usuario')
        right.pack(side='right', fill='y', padx=10, pady=10)

        def lbl_entry(parent, label, show=''):
            ttk.Label(parent, text=label).pack(padx=12, pady=(10, 2), anchor='w')
            e = ttk.Entry(parent, width=22, show=show)
            e.pack(padx=12, pady=2)
            return e

        self.usr_name = lbl_entry(right, 'Usuario:')
        self.usr_pass = lbl_entry(right, 'Contraseña:', show='*')
        ttk.Label(right, text='(dejar en blanco para no cambiar)',
                  font=('Segoe UI', 8), foreground='gray').pack(padx=12, anchor='w')

        ttk.Label(right, text='Sucursal:').pack(padx=12, pady=(10, 2), anchor='w')
        self.usr_branch = ttk.Combobox(right, state='readonly', width=20)
        self.usr_branch.pack(padx=12, pady=2)
        self._branch_combos.append(self.usr_branch)

        ttk.Separator(right).pack(fill='x', padx=12, pady=10)
        ttk.Button(right, text='➕  Agregar',   command=self._add_user).pack(padx=12, pady=4, fill='x')
        ttk.Button(right, text='✏️  Modificar', command=self._edit_user).pack(padx=12, pady=4, fill='x')
        ttk.Button(right, text='🗑  Eliminar',  command=self._del_user).pack(padx=12, pady=4, fill='x')
        ttk.Button(right, text='🔄  Limpiar',   command=self._clear_user_form).pack(padx=12, pady=4, fill='x')

        self._refresh_users()

    def _refresh_users(self):
        self.usr_tree.delete(*self.usr_tree.get_children())
        conn = get_conn()
        rows = conn.execute(
            "SELECT id, username, branch FROM users ORDER BY username"
        ).fetchall()
        conn.close()
        for r in rows:
            self.usr_tree.insert('', 'end', iid=str(r['id']),
                                 values=(r['username'], r['branch'] or ''))

    def _on_user_select(self, _event=None):
        sel = self.usr_tree.selection()
        if not sel:
            return
        vals = self.usr_tree.item(sel[0])['values']
        self.usr_name.delete(0, 'end')
        self.usr_name.insert(0, vals[0])
        self.usr_pass.delete(0, 'end')
        # Mostrar contraseña guardada en texto plano
        conn = get_conn()
        row = conn.execute(
            "SELECT password_plain FROM users WHERE id = ?", (sel[0],)
        ).fetchone()
        conn.close()
        if row and row['password_plain']:
            self.usr_pass.insert(0, row['password_plain'])
        self.usr_branch.set(vals[1] if vals[1] else '')

    def _add_user(self):
        username = self.usr_name.get().strip()
        password = self.usr_pass.get().strip()
        branch   = self.usr_branch.get().strip()
        if not username or not password:
            messagebox.showwarning("Atención", "Completá usuario y contraseña.", parent=self)
            return
        if not branch:
            messagebox.showwarning("Atención", "Asigná una sucursal al usuario.", parent=self)
            return
        conn = get_conn()
        try:
            conn.execute(
                "INSERT INTO users (username, password_hash, password_plain, branch) VALUES (?, ?, ?, ?)",
                (username, hash_password(username, password), password, branch)
            )
            conn.commit()
            messagebox.showinfo("Listo", f"Usuario '{username}' creado.", parent=self)
            self._clear_user_form()
            self._refresh_users()
        except Exception:
            messagebox.showerror("Error", f"Ya existe un usuario con ese nombre.", parent=self)
        finally:
            conn.close()

    def _edit_user(self):
        sel = self.usr_tree.selection()
        if not sel:
            messagebox.showwarning("Atención", "Seleccioná un usuario.", parent=self)
            return
        uid      = sel[0]
        username = self.usr_name.get().strip()
        password = self.usr_pass.get().strip()
        branch   = self.usr_branch.get().strip()
        if not username:
            messagebox.showwarning("Atención", "Completá el nombre de usuario.", parent=self)
            return
        conn = get_conn()
        if password:
            conn.execute(
                "UPDATE users SET username=?, password_hash=?, password_plain=?, branch=? WHERE id=?",
                (username, hash_password(username, password), password, branch or None, uid)
            )
        else:
            conn.execute(
                "UPDATE users SET username=?, branch=? WHERE id=?",
                (username, branch or None, uid)
            )
        conn.commit()
        conn.close()
        messagebox.showinfo("Listo", "Usuario modificado.", parent=self)
        self._clear_user_form()
        self._refresh_users()

    def _del_user(self):
        sel = self.usr_tree.selection()
        if not sel:
            messagebox.showwarning("Atención", "Seleccioná un usuario.", parent=self)
            return
        vals = self.usr_tree.item(sel[0])['values']
        if not messagebox.askyesno("Confirmar", f"¿Eliminar usuario '{vals[0]}'?", parent=self):
            return
        conn = get_conn()
        conn.execute("DELETE FROM users WHERE id = ?", (sel[0],))
        conn.commit()
        conn.close()
        self._refresh_users()

    def _clear_user_form(self):
        self.usr_name.delete(0, 'end')
        self.usr_pass.delete(0, 'end')
        self.usr_branch.set('')

    # ─── TAB: SINCRONIZAR ─────────────────────────────────────────────────────

    def _build_tab_sincronizar(self, nb):
        tab = ttk.Frame(nb)
        nb.add(tab, text='☁️  Sincronizar')

        # Config
        cfg_frame = ttk.LabelFrame(tab, text='Configuración de GitHub')
        cfg_frame.pack(fill='x', padx=15, pady=12)

        grid = ttk.Frame(cfg_frame)
        grid.pack(fill='x', padx=10, pady=10)

        fields = [
            ('Usuario GitHub:',  'github_user'),
            ('Repositorio:',     'github_repo'),
            ('Rama (branch):',   'github_branch'),
            ('Carpeta en repo:', 'github_path_prefix'),
        ]
        self._cfg_entries: dict[str, ttk.Entry] = {}
        for row_idx, (label, key) in enumerate(fields):
            ttk.Label(grid, text=label).grid(row=row_idx, column=0, sticky='w', padx=5, pady=3)
            e = ttk.Entry(grid, width=30)
            e.insert(0, self.cfg.get(key, ''))
            e.grid(row=row_idx, column=1, sticky='ew', padx=5, pady=3)
            self._cfg_entries[key] = e

        # Ruta del repositorio git local (para Git Push)
        repo_row = len(fields)
        ttk.Label(grid, text='Ruta repo git:').grid(row=repo_row, column=0, sticky='w', padx=5, pady=3)
        repo_frame = ttk.Frame(grid)
        repo_frame.grid(row=repo_row, column=1, sticky='ew', padx=5, pady=3)
        self.git_repo_entry = ttk.Entry(repo_frame, width=36)
        self.git_repo_entry.insert(0, self.cfg.get('git_repo_path', ''))
        self.git_repo_entry.pack(side='left', fill='x', expand=True)
        ttk.Button(
            repo_frame, text='📁',
            command=lambda: self._browse_repo_path()
        ).pack(side='left', padx=(4, 0))
        self._cfg_entries['git_repo_path'] = self.git_repo_entry

        ttk.Label(grid, text='Token GitHub:').grid(row=repo_row + 1, column=0, sticky='w', padx=5, pady=3)
        self.token_entry = ttk.Entry(grid, width=50, show='*')
        self.token_entry.insert(0, self.cfg.get('github_token', ''))
        self.token_entry.grid(row=repo_row + 1, column=1, sticky='ew', padx=5, pady=3)
        grid.columnconfigure(1, weight=1)

        ttk.Label(cfg_frame,
                  text='🔑 Generá el token en github.com/settings/tokens → Fine-grained token → Contents: Read & Write',
                  foreground='#555', font=('Segoe UI', 8)).pack(padx=10, pady=(0, 6), anchor='w')

        ttk.Button(cfg_frame, text='💾 Guardar configuración', command=self._save_cfg).pack(pady=(0, 10))

        # Sync
        sync_frame = ttk.LabelFrame(tab, text='Sincronizar con GitHub Pages')
        sync_frame.pack(fill='x', padx=15, pady=0)
        ttk.Label(sync_frame,
                  text='Sube todos los datos (sucursales, clientes, usuarios) al repositorio de GitHub.\n'
                       'Los operadores verán los cambios inmediatamente.',
                  wraplength=600).pack(padx=10, pady=8)
        btn_row = ttk.Frame(sync_frame)
        btn_row.pack(pady=(0, 10))
        ttk.Button(btn_row, text='🚀 Sincronizar ahora', command=self._sync_now).pack(side='left', padx=6)
        ttk.Button(btn_row, text='📤 Git Push', command=self._git_push).pack(side='left', padx=6)

        # Log
        log_frame = ttk.LabelFrame(tab, text='Resultado')
        log_frame.pack(fill='both', expand=True, padx=15, pady=12)
        self.sync_log = tk.Text(log_frame, height=8, state='disabled',
                                font=('Courier New', 9), bg='#1e1e1e', fg='#d4d4d4',
                                insertbackground='white')
        self.sync_log.pack(fill='both', expand=True, padx=5, pady=5)

    def _browse_repo_path(self):
        path = filedialog.askdirectory(
            title='Seleccioná la carpeta raíz del repositorio git',
            parent=self
        )
        if path:
            self.git_repo_entry.delete(0, 'end')
            self.git_repo_entry.insert(0, path)

    def _save_cfg(self):
        for key, entry in self._cfg_entries.items():
            self.cfg[key] = entry.get().strip()
        self.cfg['github_token'] = self.token_entry.get().strip()
        save_config(self.cfg)
        messagebox.showinfo("Listo", "Configuración guardada.", parent=self)

    def _log_sync(self, msg: str):
        self.sync_log.config(state='normal')
        self.sync_log.insert('end', msg + '\n')
        self.sync_log.see('end')
        self.sync_log.config(state='disabled')
        self.update_idletasks()

    def _sync_now(self):
        token = self.token_entry.get().strip()
        if not token:
            messagebox.showwarning("Atención", "Ingresá el token de GitHub primero.", parent=self)
            return

        self.sync_log.config(state='normal')
        self.sync_log.delete('1.0', 'end')
        self.sync_log.config(state='disabled')

        def run():
            try:
                syncer = GitHubSync(
                    token=token,
                    user=self.cfg.get('github_user',   'WALLE802'),
                    repo=self.cfg.get('github_repo',   'VENTAS_CLIENTES'),
                    branch=self.cfg.get('github_branch', 'main'),
                    path_prefix=self.cfg.get('github_path_prefix', ''),
                )
                self._log_sync("⏳ Conectando con GitHub...")
                syncer.sync_all(log_callback=self._log_sync)
                self._log_sync("\n✅ Sincronización completada exitosamente.")
            except Exception as e:
                self._log_sync(f"\n❌ Error: {e}")
                messagebox.showerror("Error de sincronización", str(e), parent=self)

        threading.Thread(target=run, daemon=True).start()

    def _git_push(self):
        self.sync_log.config(state='normal')
        self.sync_log.delete('1.0', 'end')
        self.sync_log.config(state='disabled')

        def run():
            import subprocess
            # Preferir la ruta configurada; si no, autodetectar
            configured = self.cfg.get('git_repo_path', '').strip()
            if configured and os.path.isdir(configured):
                repo_dir = configured
            elif getattr(sys, 'frozen', False):
                self._log_sync(
                    '❌ Error: configurá la "Ruta repo git" en la pestaña Sincronizar\n'
                    '   (carpeta raíz del repositorio clonado en esta PC)'
                )
                return
            else:
                repo_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), '..'))
            branch = self.cfg.get('github_branch', 'main')
            try:
                # 1. Fetch + reset al remote (toma los cambios remotos sin conflictos)
                self._log_sync('⬇️  Sincronizando con remote...')
                fetch = subprocess.run(
                    ['git', 'fetch', 'origin'],
                    cwd=repo_dir, capture_output=True, text=True
                )
                if fetch.returncode != 0:
                    raise RuntimeError(f"Error en git fetch:\n{fetch.stderr or fetch.stdout}")
                subprocess.run(
                    ['git', 'reset', '--soft', f'origin/{branch}'],
                    cwd=repo_dir, capture_output=True, text=True
                )
                self._log_sync('   sync OK')

                # 2. Stage y commit de los archivos del proyecto
                self._log_sync('📦 Commiteando archivos...')
                subprocess.run(
                    ['git', 'add', 'index.html', 'app.html',
                     'css/style.css', 'js/app.js', 'js/config.js',
                     'js/auth.js', 'backoffice/main.py'],
                    cwd=repo_dir, capture_output=True, text=True
                )
                result = subprocess.run(
                    ['git', 'commit', '-m', 'backoffice: actualizacion de datos y configuracion'],
                    cwd=repo_dir, capture_output=True, text=True
                )
                if result.returncode not in (0, 1):
                    raise RuntimeError(result.stderr or result.stdout)
                if 'nothing to commit' in result.stdout:
                    self._log_sync('ℹ️  Sin cambios nuevos en archivos de código.')
                else:
                    self._log_sync(result.stdout.strip())

                # 3. Push
                self._log_sync('📤 Haciendo push...')
                push = subprocess.run(
                    ['git', 'push', 'origin', branch],
                    cwd=repo_dir, capture_output=True, text=True
                )
                self._log_sync(f'   {(push.stderr or push.stdout).strip()[:300]}')
                if push.returncode != 0:
                    raise RuntimeError(push.stderr or push.stdout)
                self._log_sync('\n✅ Git push completado.')
            except Exception as e:
                self._log_sync(f'\n❌ Error: {e}')

        threading.Thread(target=run, daemon=True).start()

    # ─── TAB: MENSAJE WHATSAPP ──────────────────────────────────────────────────

    def _build_tab_mensaje(self, nb):
        tab = ttk.Frame(nb)
        nb.add(tab, text='💬  Mensaje WA')

        info = ttk.LabelFrame(tab, text='Mensaje de promoción (WhatsApp)')
        info.pack(fill='x', padx=15, pady=12)
        ttk.Label(
            info,
            text='Usá {nombre} donde quieras insertar el nombre del cliente.\n'
                 'Ejemplo: "Hola {nombre}! 👋 Tenemos una promo exclusiva para vos."',
            foreground='#555'
        ).pack(padx=10, pady=8, anchor='w')

        editor_frame = ttk.LabelFrame(tab, text='Mensaje')
        editor_frame.pack(fill='both', expand=True, padx=15, pady=0)

        self.promo_text = tk.Text(
            editor_frame, height=6, wrap='word',
            font=('Segoe UI', 11), padx=8, pady=8
        )
        sb = ttk.Scrollbar(editor_frame, command=self.promo_text.yview)
        self.promo_text.configure(yscrollcommand=sb.set)
        self.promo_text.pack(side='left', fill='both', expand=True, padx=(5, 0), pady=5)
        sb.pack(side='right', fill='y', pady=5)
        self.promo_text.bind('<KeyRelease>', lambda _: self._update_promo_preview())

        preview_frame = ttk.LabelFrame(tab, text='Vista previa  (nombre de ejemplo: "María")')
        preview_frame.pack(fill='x', padx=15, pady=10)
        self.promo_preview = tk.Label(
            preview_frame, text='', wraplength=700, justify='left',
            font=('Segoe UI', 10), foreground='#1a73e8'
        )
        self.promo_preview.pack(padx=10, pady=8, anchor='w')

        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill='x', padx=15, pady=(0, 12))
        ttk.Button(btn_frame, text='💾  Guardar mensaje',
                   command=self._save_promo_msg).pack(side='left')
        ttk.Label(
            btn_frame,
            text='  ⚠️  Después de guardar, sincronizá para publicar los cambios.',
            foreground='#e65100', font=('Segoe UI', 9)
        ).pack(side='left', padx=10)

        self._load_promo_msg()

    def _config_js_path(self) -> str:
        if getattr(sys, 'frozen', False):
            return os.path.join(os.path.dirname(sys.executable), 'js', 'config.js')
        return os.path.normpath(
            os.path.join(os.path.dirname(__file__), '..', 'js', 'config.js')
        )

    def _load_promo_msg(self):
        try:
            content = open(self._config_js_path(), encoding='utf-8').read()
            # Coincide con PROMO_MSG: "..." o '...' incluyendo saltos de línea reales
            m = re.search(r'PROMO_MSG:\s*"((?:[^"\\]|\\.)*)"', content, re.DOTALL)
            if not m:
                m = re.search(r"PROMO_MSG:\s*'((?:[^'\\]|\\.)*)'", content, re.DOTALL)
            if m:
                msg = (m.group(1)
                       .replace('\\n', '\n')
                       .replace('\\"', '"')
                       .replace("\\'", "'")
                       .replace('\\\\', '\\'))
                self.promo_text.delete('1.0', 'end')
                self.promo_text.insert('1.0', msg)
                self._update_promo_preview()
        except Exception:
            pass

    def _update_promo_preview(self):
        msg = self.promo_text.get('1.0', 'end-1c')
        self.promo_preview.config(text=msg.replace('{nombre}', 'María'))

    def _save_promo_msg(self):
        msg = self.promo_text.get('1.0', 'end-1c').strip()
        if not msg:
            messagebox.showwarning('Atención', 'El mensaje no puede estar vacío.', parent=self)
            return
        path = self._config_js_path()
        try:
            with open(path, 'r', encoding='utf-8') as f:
                content = f.read()
            # Escapar para JS: backslashes, comillas y saltos de línea
            escaped = (msg
                       .replace('\\', '\\\\')
                       .replace('"', '\\"')
                       .replace('\r\n', '\\n')
                       .replace('\n', '\\n')
                       .replace('\r', '\\n'))
            # Usar anchors fijos para reemplazar todo el bloque PROMO_MSG
            start = content.find('    PROMO_MSG:')
            end   = content.find('    get RAW_BASE')
            if start == -1 or end == -1:
                raise ValueError('No se encontró PROMO_MSG en config.js')
            new_content = content[:start] + f'    PROMO_MSG: "{escaped}",\n\n' + content[end:]
            with open(path, 'w', encoding='utf-8') as f:
                f.write(new_content)
            messagebox.showinfo(
                'Listo',
                '✅ Mensaje guardado en config.js\n\n'
                'Sincronizá los datos (☁️ Sincronizar) y luego hacé\n'
                'git push para que el cambio se vea en la página web.',
                parent=self
            )
        except Exception as e:
            messagebox.showerror('Error', str(e), parent=self)

    # ─── TAB: REPORTES ────────────────────────────────────────────────────────

    def _build_tab_reportes(self, nb):
        tab = ttk.Frame(nb)
        nb.add(tab, text='📊  Reportes')

        # Fila superior: filtros
        top = ttk.Frame(tab)
        top.pack(fill='x', padx=10, pady=6)

        ttk.Label(top, text='Desde:').pack(side='left')
        self.rpt_date_from = ttk.Entry(top, width=12)
        self.rpt_date_from.insert(0, date.today().strftime('%Y-%m-%d'))
        self.rpt_date_from.pack(side='left', padx=(4, 6))

        ttk.Label(top, text='Hasta:').pack(side='left')
        self.rpt_date_to = ttk.Entry(top, width=12)
        self.rpt_date_to.insert(0, date.today().strftime('%Y-%m-%d'))
        self.rpt_date_to.pack(side='left', padx=(4, 10))

        ttk.Label(top, text='Usuario:').pack(side='left')
        self.rpt_user_var = tk.StringVar()
        self.rpt_user_cb = ttk.Combobox(top, textvariable=self.rpt_user_var,
                                        state='readonly', width=14)
        self.rpt_user_cb.pack(side='left', padx=(4, 10))

        ttk.Label(top, text='Sucursal:').pack(side='left')
        self.rpt_branch_var = tk.StringVar()
        self.rpt_branch_cb = ttk.Combobox(top, textvariable=self.rpt_branch_var,
                                          state='readonly', width=14)
        self.rpt_branch_cb.pack(side='left', padx=(4, 10))
        self._branch_combos.append(self.rpt_branch_cb)

        # Fila de botones
        btn_row = ttk.Frame(tab)
        btn_row.pack(fill='x', padx=10, pady=(0, 4))
        ttk.Button(btn_row, text='🔍 Ver registros',
                   command=self._load_report).pack(side='left', padx=3)
        ttk.Button(btn_row, text='☁️ Importar de GitHub',
                   command=self._import_logs_from_github).pack(side='left', padx=3)
        ttk.Button(btn_row, text='� Exportar Excel',
                   command=self._export_excel).pack(side='left', padx=3)

        # Tarjetas de resumen (ocultas hasta cargar datos)
        self.rpt_stats_frame = ttk.LabelFrame(tab, text='Resumen del período')
        self._stat_labels: dict[str, tk.StringVar] = {}
        stat_defs = [
            ('total',    'Total gestiones'),
            ('llamada',  '📞 Llamadas'),
            ('whatsapp', '💬 WhatsApp'),
            ('sms',      '✉️ SMS'),
            ('promo',    '🎁 Promos'),
            ('usuarios', '👤 Usuarios activos'),
            ('clientes', '🧑 Clientes únicos'),
        ]
        for key, label in stat_defs:
            var = tk.StringVar(value='—')
            self._stat_labels[key] = var
            card = ttk.Frame(self.rpt_stats_frame, relief='groove', borderwidth=1)
            card.pack(side='left', padx=6, pady=6, ipadx=10, ipady=4)
            ttk.Label(card, text=label, font=('Segoe UI', 8)).pack()
            ttk.Label(card, textvariable=var, font=('Segoe UI', 18, 'bold'),
                      foreground='#1a73e8').pack()

        # Barra de progreso de importación (oculta por defecto)
        self.rpt_progress_frame = ttk.Frame(tab)
        self.rpt_progress = ttk.Progressbar(self.rpt_progress_frame, mode='determinate')
        self.rpt_progress.pack(fill='x', padx=5, pady=2)
        self.rpt_progress_lbl = ttk.Label(self.rpt_progress_frame, text='')
        self.rpt_progress_lbl.pack()

        # Tabla
        frame = ttk.Frame(tab)
        frame.pack(fill='both', expand=True, padx=10, pady=(0, 5))

        cols = ('Fecha', 'Hora', 'Usuario', 'Sucursal', 'DNI', 'Nombre', 'Tipo', 'Teléfono')
        self.rpt_tree = ttk.Treeview(frame, columns=cols, show='headings')
        widths = (90, 55, 100, 90, 90, 180, 90, 120)
        for col, w in zip(cols, widths):
            self.rpt_tree.heading(col, text=col)
            self.rpt_tree.column(col, width=w, minwidth=50)

        vsb = ttk.Scrollbar(frame, orient='vertical',   command=self.rpt_tree.yview)
        hsb = ttk.Scrollbar(frame, orient='horizontal', command=self.rpt_tree.xview)
        self.rpt_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.rpt_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        self.rpt_status = ttk.Label(tab, text='')
        self.rpt_status.pack(pady=2)
        self._rpt_data: list[dict] = []

        # Cargar usuarios disponibles
        self._refresh_rpt_users()

    def _refresh_rpt_users(self):
        conn = get_conn()
        users = [r['username'] for r in conn.execute(
            "SELECT DISTINCT username FROM contacts ORDER BY username"
        ).fetchall()]
        conn.close()
        self.rpt_user_cb['values'] = ['(todos)'] + users
        self.rpt_user_cb.set('(todos)')
        current_branch = self.rpt_branch_var.get()
        if not current_branch:
            self.rpt_branch_cb.set('(todas)')

    def _load_report(self):
        date_from = self.rpt_date_from.get().strip()
        date_to   = self.rpt_date_to.get().strip()
        username  = self.rpt_user_var.get()
        branch    = self.rpt_branch_var.get()

        self.rpt_tree.delete(*self.rpt_tree.get_children())
        self._rpt_data = []

        conn = get_conn()
        query  = "SELECT date, time, username, branch, client_dni, client_name, contact_type, phone_used FROM contacts WHERE 1=1"
        params = []
        if date_from:
            query += " AND date >= ?"
            params.append(date_from)
        if date_to:
            query += " AND date <= ?"
            params.append(date_to)
        if username and username != '(todos)':
            query += " AND username = ?"
            params.append(username)
        if branch and branch not in ('(todas)', ''):
            query += " AND branch = ?"
            params.append(branch)
        query += " ORDER BY date DESC, time DESC"

        rows = conn.execute(query, params).fetchall()
        conn.close()

        for r in rows:
            entry = {
                'date':         r['date'],
                'time':         r['time'],
                'username':     r['username'],
                'branch':       r['branch'],
                'client_dni':   r['client_dni']   or '',
                'client_name':  r['client_name']  or '',
                'contact_type': r['contact_type'],
                'phone_used':   r['phone_used']   or '',
            }
            self._rpt_data.append(entry)
            self.rpt_tree.insert('', 'end', values=(
                r['date'], r['time'], r['username'], r['branch'],
                r['client_dni'] or '', r['client_name'] or '',
                r['contact_type'], r['phone_used'] or '',
            ))

        self.rpt_status.config(text=f"{len(rows)} registro(s) encontrado(s)")
        self._update_stats()

    def _update_stats(self):
        from collections import Counter
        data = self._rpt_data
        if not data:
            for var in self._stat_labels.values():
                var.set('0')
            self.rpt_stats_frame.pack_forget()
            return

        tipos = Counter(e['contact_type'] for e in data)
        promos = sum(v for k, v in tipos.items() if 'promo' in k.lower())

        self._stat_labels['total'].set(str(len(data)))
        self._stat_labels['llamada'].set(str(tipos.get('llamada', 0)))
        self._stat_labels['whatsapp'].set(str(tipos.get('whatsapp', 0)))
        self._stat_labels['sms'].set(str(tipos.get('sms', 0)))
        self._stat_labels['promo'].set(str(promos))
        self._stat_labels['usuarios'].set(str(len(set(e['username'] for e in data))))
        self._stat_labels['clientes'].set(str(len(set(
            e['client_dni'] for e in data if e['client_dni']
        ))))

        # Mostrar el frame de stats justo antes de la barra de progreso
        self.rpt_stats_frame.pack(fill='x', padx=10, pady=(0, 4),
                                   before=self.rpt_progress_frame)

    def _import_logs_from_github(self):
        token = self.cfg.get('github_token', '') or self.token_entry.get().strip()
        if not token:
            messagebox.showwarning("Atención",
                                   "Configurá el token de GitHub en la pestaña Sincronizar.", parent=self)
            return

        self.rpt_progress_frame.pack(fill='x', padx=10, pady=2, before=self.rpt_tree.master)

        def run():
            try:
                syncer = GitHubSync(
                    token=token,
                    user=self.cfg.get('github_user',   'WALLE802'),
                    repo=self.cfg.get('github_repo',   'VENTAS_CLIENTES'),
                    branch=self.cfg.get('github_branch', 'main'),
                    path_prefix=self.cfg.get('github_path_prefix', ''),
                )
                self.rpt_progress_lbl.config(text='Obteniendo fechas disponibles...')
                dates = syncer.list_log_dates()
                if not dates:
                    self.after(0, lambda: messagebox.showinfo(
                        "Sin datos", "No hay logs en el repositorio.", parent=self))
                    return

                total      = len(dates)
                imported   = 0
                duplicates = 0
                conn = get_conn()
                for i, d in enumerate(dates, 1):
                    self.rpt_progress['value'] = (i / total) * 100
                    self.rpt_progress_lbl.config(text=f'Importando {d}  ({i}/{total})')
                    self.update_idletasks()
                    try:
                        logs = syncer.get_logs(d)
                    except Exception:
                        continue
                    for e in logs:
                        try:
                            conn.execute(
                                """INSERT OR IGNORE INTO contacts
                                   (date, time, username, branch, client_dni,
                                    client_name, contact_type, phone_used)
                                   VALUES (?,?,?,?,?,?,?,?)""",
                                (
                                    e.get('date', d),
                                    e.get('time', ''),
                                    e.get('username', ''),
                                    e.get('branch', ''),
                                    e.get('client_dni', ''),
                                    e.get('client_name', ''),
                                    e.get('contact_type', ''),
                                    e.get('phone_used', ''),
                                )
                            )
                            if conn.execute("SELECT changes()").fetchone()[0]:
                                imported += 1
                            else:
                                duplicates += 1
                        except Exception:
                            pass
                conn.commit()
                conn.close()

                self.rpt_progress_lbl.config(text='¡Importación completada!')
                self.after(200, lambda: self.rpt_progress_frame.pack_forget())
                self._refresh_rpt_users()
                self._load_report()
                self.after(0, lambda: messagebox.showinfo(
                    "Importación completada",
                    f"✅ {imported} registro(s) nuevos importados.\n"
                    f"   {duplicates} duplicado(s) omitidos.",
                    parent=self
                ))
            except Exception as e:
                err_msg = str(e)
                self.after(200, lambda: self.rpt_progress_frame.pack_forget())
                self.after(0, lambda m=err_msg: messagebox.showerror("Error", m, parent=self))

        threading.Thread(target=run, daemon=True).start()

    def _export_excel(self):
        if not self._rpt_data:
            messagebox.showwarning("Atención", "No hay datos para exportar.", parent=self)
            return

        from collections import Counter
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        date_label = self.rpt_date_from.get().strip()
        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel', '*.xlsx')],
            initialfile=f"reporte_{date_label}.xlsx",
            parent=self
        )
        if not path:
            return

        wb = openpyxl.Workbook()

        # ── Hoja 1: Resumen ──────────────────────────────────────────────────
        ws_r = wb.active  # type: ignore[assignment]
        assert ws_r is not None
        ws_r.title = 'Resumen'

        header_fill = PatternFill('solid', fgColor='1a73e8')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        bold        = Font(bold=True)
        center      = Alignment(horizontal='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def hdr(ws, row, col, text):
            c = ws.cell(row=row, column=col, value=text)
            c.font = header_font; c.fill = header_fill
            c.alignment = center; c.border = thin_border

        def cell(ws, row, col, value):
            c = ws.cell(row=row, column=col, value=value)
            c.border = thin_border; c.alignment = center

        data = self._rpt_data
        tipos = Counter(e['contact_type'] for e in data)
        usuarios = Counter(e['username'] for e in data)
        sucursales = Counter(e['branch'] for e in data)

        # Bloque: totales por tipo
        hdr(ws_r, 1, 1, 'Tipo de gestión');  hdr(ws_r, 1, 2, 'Cantidad')
        for i, (tipo, cnt) in enumerate(sorted(tipos.items()), start=2):
            cell(ws_r, i, 1, tipo); cell(ws_r, i, 2, cnt)
        row_sep = len(tipos) + 3

        # Bloque: totales por usuario
        hdr(ws_r, row_sep, 1, 'Usuario');    hdr(ws_r, row_sep, 2, 'Gestiones')
        for i, (usr, cnt) in enumerate(sorted(usuarios.items(), key=lambda x: -x[1]), start=1):
            cell(ws_r, row_sep + i, 1, usr); cell(ws_r, row_sep + i, 2, cnt)
        row_sep2 = row_sep + len(usuarios) + 2

        # Bloque: totales por sucursal
        hdr(ws_r, row_sep2, 1, 'Sucursal');  hdr(ws_r, row_sep2, 2, 'Gestiones')
        for i, (suc, cnt) in enumerate(sorted(sucursales.items(), key=lambda x: -x[1]), start=1):
            cell(ws_r, row_sep2 + i, 1, suc); cell(ws_r, row_sep2 + i, 2, cnt)

        # Total general
        ws_r.cell(row=1, column=4, value='Total gestiones').font = bold
        ws_r.cell(row=2, column=4, value=len(data)).font = Font(bold=True, size=16, color='1a73e8')
        ws_r.cell(row=3, column=4, value='Usuarios activos').font = bold
        ws_r.cell(row=4, column=4, value=len(usuarios)).font = Font(bold=True, size=14)
        ws_r.cell(row=5, column=4, value='Clientes únicos').font = bold
        ws_r.cell(row=6, column=4, value=len(set(
            e['client_dni'] for e in data if e['client_dni']
        ))).font = Font(bold=True, size=14)

        ws_r.column_dimensions['A'].width = 22
        ws_r.column_dimensions['B'].width = 14
        ws_r.column_dimensions['D'].width = 18

        # ── Hoja 2: Detalle ──────────────────────────────────────────────────
        ws_d = wb.create_sheet('Detalle')
        headers = ['Fecha', 'Hora', 'Usuario', 'Sucursal', 'DNI',
                   'Nombre cliente', 'Tipo gestión', 'Teléfono']
        col_widths = [12, 8, 18, 14, 14, 28, 14, 16]
        for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
            c = ws_d.cell(row=1, column=col, value=h)
            c.font = header_font; c.fill = header_fill
            c.alignment = center; c.border = thin_border
            ws_d.column_dimensions[get_column_letter(col)].width = w

        for row_i, e in enumerate(data, start=2):
            for col_i, key in enumerate(
                ('date', 'time', 'username', 'branch',
                 'client_dni', 'client_name', 'contact_type', 'phone_used'),
                start=1
            ):
                cell(ws_d, row_i, col_i, e.get(key, ''))

        ws_d.freeze_panes = 'A2'
        ws_d.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        wb.save(path)
        messagebox.showinfo("Listo", f"Excel exportado:\n{path}", parent=self)

    def _export_csv(self):
        if not self._rpt_data:
            messagebox.showwarning("Atención", "No hay datos para exportar.", parent=self)
            return
        date_label = self.rpt_date_from.get().strip()
        path = filedialog.asksaveasfilename(
            defaultextension='.csv',
            filetypes=[('CSV', '*.csv')],
            initialfile=f"reporte_{date_label}.csv",
            parent=self
        )
        if not path:
            return
        fields = ['date', 'time', 'username', 'branch', 'client_dni',
                  'client_name', 'contact_type', 'phone_used']
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(self._rpt_data)
        messagebox.showinfo("Listo", f"Exportado:\n{path}", parent=self)


# ─── Entry point ──────────────────────────────────────────────────────────────

if __name__ == '__main__':
    app = BackofficeApp()
    app.mainloop()
